"""
XEOSIterativeSegmentation — 3D Slicer Extension
================================================
PET lesion iterative segmentation using Pedro's plateau-detection method.

Workflow:
  1. Load DICOM study via Slicer's built-in DICOM browser (no conversion needed)
  2. Select the PET and CT volumes from the loaded series
  3. Auto-generate an Otsu-based initial segmentation → displayed in the viewer
  4. User corrects the segmentation interactively (add/remove lesion labels)
  5. Click "Run Iterative Segmentation" to detect the activity plateau per lesion
  6. View metrics and export them to Excel

Author:  Yazdan Salimi  (salimiyazdan@gmail.com)  — ported to Slicer by Claude
"""

import os
import numpy as np
import vtk
import qt
import ctk
import slicer
from slicer.ScriptedLoadableModule import (
    ScriptedLoadableModule,
    ScriptedLoadableModuleWidget,
    ScriptedLoadableModuleLogic,
    ScriptedLoadableModuleTest,
)
from slicer.util import VTKObservationMixin


# ─────────────────────────────────────────────────────────────────────────────
# Module descriptor
# ─────────────────────────────────────────────────────────────────────────────
class XEOSIterativeSegmentation(ScriptedLoadableModule):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent.title = "XEOS Iterative Segmentation"
        self.parent.categories = ["Quantification", "PET"]
        self.parent.dependencies = []
        self.parent.contributors = ["Yazdan Salimi"]
        self.parent.helpText = (
            "Iterative PET lesion segmentation using plateau detection "
            "(Pedro methodology). Loads DICOM directly in Slicer, generates "
            "an Otsu initial segmentation that the user can correct, then runs "
            "iterative thresholding and exports metrics to Excel."
        )
        self.parent.acknowledgementText = (
            "Developed for XEOS PET/CT scanner studies."
        )


# ─────────────────────────────────────────────────────────────────────────────
# Widget  (GUI)
# ─────────────────────────────────────────────────────────────────────────────
class XEOSIterativeSegmentationWidget(ScriptedLoadableModuleWidget, VTKObservationMixin):

    def __init__(self, parent=None):
        ScriptedLoadableModuleWidget.__init__(self, parent)
        VTKObservationMixin.__init__(self)
        self.logic = None
        self._parameterNode = None

    # ------------------------------------------------------------------
    def setup(self):
        super().setup()
        self.logic = XEOSIterativeSegmentationLogic()

        # ── 1. DICOM loader hint ───────────────────────────────────────
        dicomBox = ctk.ctkCollapsibleButton()
        dicomBox.text = "Step 1 – Load DICOM"
        self.layout.addWidget(dicomBox)
        dicomForm = qt.QFormLayout(dicomBox)

        loadDicomBtn = qt.QPushButton("🗂  Open DICOM Browser")
        loadDicomBtn.toolTip = (
            "Open Slicer's built-in DICOM browser to import and load your study."
        )
        loadDicomBtn.setStyleSheet("""
            QPushButton {
                background-color: #1565C0;
                color: white;
                font-weight: bold;
                font-size: 13px;
                padding: 6px 12px;
                border-radius: 5px;
                border: none;
            }
            QPushButton:hover { background-color: #1976D2; }
            QPushButton:pressed { background-color: #0D47A1; }
        """)
        loadDicomBtn.clicked.connect(self.onOpenDicomBrowser)
        dicomForm.addRow(loadDicomBtn)

        # ── 2. Volume selection ────────────────────────────────────────
        volBox = ctk.ctkCollapsibleButton()
        volBox.text = "Step 2 – Select Volumes"
        self.layout.addWidget(volBox)
        volForm = qt.QFormLayout(volBox)

        self.petSelector = slicer.qMRMLNodeComboBox()
        self.petSelector.nodeTypes = ["vtkMRMLScalarVolumeNode"]
        self.petSelector.selectNodeUponCreation = False
        self.petSelector.addEnabled = False
        self.petSelector.removeEnabled = False
        self.petSelector.noneEnabled = True
        self.petSelector.showHidden = False
        self.petSelector.showChildNodeTypes = False
        self.petSelector.setMRMLScene(slicer.mrmlScene)
        self.petSelector.setToolTip("Select the PET volume (Bq/mL or SUV)")
        self.petSelector.currentNodeChanged.connect(self.onPetVolumeSelected)
        volForm.addRow("PET volume:", self.petSelector)

        self.petColormapCombo = qt.QComboBox()
        self.petColormapCombo.setToolTip("Choose the colormap applied to the PET volume in all slice views.")
        for label, name in [
            ("PET-DICOM (default)", "PET-DICOM"),
            ("PET-Heat",            "PET-Heat"),
            ("PET (rainbow)",       "PET"),
            ("Hot",                 "Hot"),
            ("Grey",                "Grey"),
        ]:
            self.petColormapCombo.addItem(label, name)
        self.petColormapCombo.currentIndexChanged.connect(self.onColormapChanged)
        volForm.addRow("PET colormap:", self.petColormapCombo)

        self.ctSelector = slicer.qMRMLNodeComboBox()
        self.ctSelector.nodeTypes = ["vtkMRMLScalarVolumeNode"]
        self.ctSelector.selectNodeUponCreation = False
        self.ctSelector.addEnabled = False
        self.ctSelector.removeEnabled = False
        self.ctSelector.noneEnabled = True
        self.ctSelector.showHidden = False
        self.ctSelector.showChildNodeTypes = False
        self.ctSelector.setMRMLScene(slicer.mrmlScene)
        self.ctSelector.setToolTip("Select the CT volume (optional, for display only)")
        self.ctSelector.currentNodeChanged.connect(self.onCtVolumeSelected)
        volForm.addRow("CT volume (optional):", self.ctSelector)

        # ── 3. SUV / Bq/mL toggle ─────────────────────────────────────
        unitBox = ctk.ctkCollapsibleButton()
        unitBox.text = "Step 3 – Units & Parameters"
        self.layout.addWidget(unitBox)
        unitForm = qt.QFormLayout(unitBox)

        self.useSUVCheckbox = qt.QCheckBox("PET values are already in SUV")
        self.useSUVCheckbox.checked = True
        self.useSUVCheckbox.toolTip = (
            "Check if the volume is in SUV. "
            "Leave unchecked for Bq/mL (will be converted to kBq/mL internally)."
        )
        unitForm.addRow(self.useSUVCheckbox)

        self.plateauThreshSpin = qt.QDoubleSpinBox()
        self.plateauThreshSpin.minimum = 0.1
        self.plateauThreshSpin.maximum = 20.0
        self.plateauThreshSpin.singleStep = 0.5
        self.plateauThreshSpin.value = 2.0
        self.plateauThreshSpin.toolTip = (
            "Volume-change percentage below which a plateau is considered reached."
        )
        unitForm.addRow("Plateau threshold (%):", self.plateauThreshSpin)

        self.postPlateauSpin = qt.QSpinBox()
        self.postPlateauSpin.minimum = 1
        self.postPlateauSpin.maximum = 50
        self.postPlateauSpin.value = 10
        self.postPlateauSpin.toolTip = (
            "Number of extra iterations after plateau detection before stopping."
        )
        unitForm.addRow("Post-plateau iterations:", self.postPlateauSpin)

        self.iterRangeSpin = qt.QSpinBox()
        self.iterRangeSpin.minimum = 10
        self.iterRangeSpin.maximum = 2000
        self.iterRangeSpin.value = 500
        unitForm.addRow("Max iterations:", self.iterRangeSpin)

        self.thresholdStepSpin = qt.QDoubleSpinBox()
        self.thresholdStepSpin.minimum = 0.001
        self.thresholdStepSpin.maximum = 10.0
        self.thresholdStepSpin.singleStep = 0.05
        self.thresholdStepSpin.decimals = 3
        self.thresholdStepSpin.value = 0.1
        self.thresholdStepSpin.toolTip = (
            "Step size between successive threshold values (in SUV or kBq/mL). "
            "Smaller = finer resolution but more iterations. "
            "Typical range: 0.05–0.5 for SUV, 0.5–5.0 for kBq/mL."
        )
        unitForm.addRow("Threshold step size:", self.thresholdStepSpin)

        self.lowestVolSpin = qt.QDoubleSpinBox()
        self.lowestVolSpin.minimum = 0.0
        self.lowestVolSpin.maximum = 100.0
        self.lowestVolSpin.singleStep = 0.1
        self.lowestVolSpin.value = 0.0
        self.lowestVolSpin.toolTip = "Stop iterating when volume drops below this (mL)."
        unitForm.addRow("Minimum volume (mL):", self.lowestVolSpin)

        self.keepLargestCheckbox = qt.QCheckBox("Keep only largest object per lesion")
        self.keepLargestCheckbox.checked = True
        unitForm.addRow(self.keepLargestCheckbox)

        self.maskToRoughCheckbox = qt.QCheckBox("Constrain result to initial rough segment")
        self.maskToRoughCheckbox.checked = True
        unitForm.addRow(self.maskToRoughCheckbox)

        # ── 4. Initial Otsu segmentation ──────────────────────────────
        otsuBox = ctk.ctkCollapsibleButton()
        otsuBox.text = "Step 4 – Generate / Correct Initial Segmentation"
        self.layout.addWidget(otsuBox)
        otsuForm = qt.QFormLayout(otsuBox)

        self.minObjVolSpin = qt.QDoubleSpinBox()
        self.minObjVolSpin.minimum = 0.0
        self.minObjVolSpin.maximum = 10.0
        self.minObjVolSpin.singleStep = 0.05
        self.minObjVolSpin.value = 0.1
        self.minObjVolSpin.toolTip = "Remove connected components smaller than this (mL) from the Otsu result."
        otsuForm.addRow("Min object volume for Otsu (mL):", self.minObjVolSpin)

        generateOtsuBtn = qt.QPushButton("⚙  Generate Otsu Initial Segmentation")
        generateOtsuBtn.toolTip = (
            "Runs Otsu thresholding on the PET, removes small objects, dilates, "
            "labels connected components and loads result as a Segmentation node "
            "so you can edit it with the Segment Editor."
        )
        generateOtsuBtn.setStyleSheet("""
            QPushButton {
                background-color: #00695C;
                color: white;
                font-weight: bold;
                font-size: 13px;
                padding: 6px 12px;
                border-radius: 5px;
                border: none;
            }
            QPushButton:hover { background-color: #00796B; }
            QPushButton:pressed { background-color: #004D40; }
        """)
        generateOtsuBtn.clicked.connect(self.onGenerateOtsu)
        otsuForm.addRow(generateOtsuBtn)

        self.segNodeSelector = slicer.qMRMLNodeComboBox()
        self.segNodeSelector.nodeTypes = ["vtkMRMLSegmentationNode"]
        self.segNodeSelector.selectNodeUponCreation = False
        self.segNodeSelector.addEnabled = False
        self.segNodeSelector.removeEnabled = False
        self.segNodeSelector.noneEnabled = True
        self.segNodeSelector.showHidden = False
        self.segNodeSelector.setMRMLScene(slicer.mrmlScene)
        self.segNodeSelector.setToolTip(
            "After editing in Segment Editor, select the corrected segmentation here."
        )
        otsuForm.addRow("Segmentation to use:", self.segNodeSelector)

        openSegEditorBtn = qt.QPushButton("Open Segment Editor")
        openSegEditorBtn.toolTip = "Switch to the Segment Editor module to correct the segmentation."
        openSegEditorBtn.clicked.connect(self.onOpenSegmentEditor)
        otsuForm.addRow(openSegEditorBtn)

        # ── 5. Run ────────────────────────────────────────────────────
        runBox = ctk.ctkCollapsibleButton()
        runBox.text = "Step 5 – Run Iterative Segmentation"
        self.layout.addWidget(runBox)
        runForm = qt.QFormLayout(runBox)

        self.runBtn = qt.QPushButton("▶  Run Iterative Segmentation")
        self.runBtn.toolTip = "Run plateau-detection segmentation for every lesion label."
        self.runBtn.setStyleSheet("""
            QPushButton {
                background-color: #2E7D32;
                color: white;
                font-weight: bold;
                font-size: 14px;
                padding: 8px 16px;
                border-radius: 6px;
                border: none;
            }
            QPushButton:hover   { background-color: #388E3C; }
            QPushButton:pressed { background-color: #1B5E20; }
            QPushButton:disabled { background-color: #555; color: #aaa; }
        """)
        self.runBtn.clicked.connect(self.onRunSegmentation)
        runForm.addRow(self.runBtn)

        self.progressBar = qt.QProgressBar()
        self.progressBar.setRange(0, 100)
        self.progressBar.value = 0
        runForm.addRow(self.progressBar)

        self.statusLabel = qt.QLabel("Ready.")
        runForm.addRow(self.statusLabel)

        self.toggleInitSegBtn = qt.QPushButton("👁  Show Initial Segmentation")
        self.toggleInitSegBtn.toolTip = (
            "Toggle visibility of the initial Otsu/corrected segmentation. "
            "It is hidden automatically when iterative segmentation finishes."
        )
        self.toggleInitSegBtn.setEnabled(False)
        self.toggleInitSegBtn.clicked.connect(self.onToggleInitialSegmentation)
        runForm.addRow(self.toggleInitSegBtn)

        # ── 5b. Interactive plateau curve ───────────────────────────────
        curveBox = ctk.ctkCollapsibleButton()
        curveBox.text = "Step 5b – Interactive Plateau Curve (optional manual override)"
        self.layout.addWidget(curveBox)
        curveForm = qt.QFormLayout(curveBox)

        self.curveLesionSelector = qt.QComboBox()
        self.curveLesionSelector.toolTip = "Choose which lesion's volume-vs-threshold curve to inspect."
        self.curveLesionSelector.currentIndexChanged.connect(self.onCurveLesionChanged)
        curveForm.addRow("Lesion:", self.curveLesionSelector)

        # ctkVTKChartView — a plain embedded VTK chart, no layout/view-node
        # coupling.  Ships with every Slicer build.
        self.chartView = ctk.ctkVTKChartView()
        self.chartView.setMinimumHeight(300)
        self.chartView.setToolTip(
            "Click or drag on the curve to set the threshold directly. "
            "The segmentation updates live."
        )
        curveForm.addRow(self.chartView)

        # Install mouse observers for click-and-drag threshold selection.
        # We do this once here; _buildPlotForLesion keeps a reference to the
        # curve data so the callback knows which X values are valid.
        self._isDraggingChart = False
        self._chartObserverTags = []
        self._installChartMouseObservers()

        sliderRow = qt.QHBoxLayout()
        self.thresholdSlider = ctk.ctkSliderWidget()
        self.thresholdSlider.decimals = 4
        self.thresholdSlider.singleStep = 0.001
        self.thresholdSlider.setEnabled(False)
        self.thresholdSlider.toolTip = (
            "Drag to choose a threshold along the curve. The segmentation "
            "preview updates live in the slice/3D views."
        )
        self.thresholdSlider.valueChanged.connect(self.onThresholdSliderChanged)
        sliderRow.addWidget(self.thresholdSlider)
        curveForm.addRow("Threshold:", sliderRow)

        statsRow = qt.QHBoxLayout()
        self.curveVolumeLabel = qt.QLabel("Volume: —")
        self.curveMeanLabel = qt.QLabel("Mean: —")
        statsRow.addWidget(self.curveVolumeLabel)
        statsRow.addWidget(self.curveMeanLabel)
        curveForm.addRow(statsRow)

        resetAutoBtn = qt.QPushButton("Reset to Auto-Detected Plateau")
        resetAutoBtn.toolTip = "Jump the slider back to the automatically detected plateau threshold."
        resetAutoBtn.clicked.connect(self.onResetToAutoPlateau)
        curveForm.addRow(resetAutoBtn)

        self.applyThresholdBtn = qt.QPushButton("✔  Apply Selected Threshold for This Lesion")
        self.applyThresholdBtn.toolTip = (
            "Commit the currently selected threshold as this lesion's final "
            "result (updates the metrics table used for Excel export)."
        )
        self.applyThresholdBtn.setEnabled(False)
        self.applyThresholdBtn.clicked.connect(self.onApplyThreshold)
        curveForm.addRow(self.applyThresholdBtn)

        self.curveStatusLabel = qt.QLabel(
            "Run the iterative segmentation first to enable the curve."
        )
        self.curveStatusLabel.setWordWrap(True)
        curveForm.addRow(self.curveStatusLabel)

        # ── 6. Export metrics ──────────────────────────────────────────
        exportBox = ctk.ctkCollapsibleButton()
        exportBox.text = "Step 6 – Export Metrics to Excel"
        self.layout.addWidget(exportBox)
        exportForm = qt.QFormLayout(exportBox)

        # ── shared output folder (used by Steps 6 AND 7) ──────────────
        self.outputDirBtn = ctk.ctkDirectoryButton()
        self.outputDirBtn.caption = "Select output folder (shared by all exports)"
        exportForm.addRow("Output folder:", self.outputDirBtn)

        exportBtn = qt.QPushButton("💾  Save Metrics to Excel")
        exportBtn.clicked.connect(self.onExportExcel)
        exportForm.addRow(exportBtn)

        self.saveConvergencePlotCheck = qt.QCheckBox("Also save convergence plot (all lesions) as PNG")
        self.saveConvergencePlotCheck.checked = True
        self.saveConvergencePlotCheck.toolTip = (
            "Saves a single PNG showing all lesion convergence curves with "
            "their final thresholds marked. Stars = manually overridden, circles = auto."
        )
        exportForm.addRow(self.saveConvergencePlotCheck)

        showTableBtn = qt.QPushButton("📊  Show Results Table in Slicer")
        showTableBtn.toolTip = "Push the summary metrics into a Slicer Table node and open the Tables view — no saving required."
        showTableBtn.clicked.connect(self.onShowResultsTable)
        exportForm.addRow(showTableBtn)

        # ── 7. Export images & segmentations ──────────────────────────
        imgExportBox = ctk.ctkCollapsibleButton()
        imgExportBox.text = "Step 7 – Export Images & Segmentations"
        self.layout.addWidget(imgExportBox)
        imgExportForm = qt.QFormLayout(imgExportBox)

        # Note pointing to shared folder
        sharedFolderNote = qt.QLabel("📂 Uses the same output folder as Step 6 above.")
        sharedFolderNote.setStyleSheet("color: #888; font-style: italic; font-size: 11px;")
        imgExportForm.addRow(sharedFolderNote)

        # Format selector
        self.exportFormatCombo = qt.QComboBox()
        self.exportFormatCombo.addItem("NRRD (.nrrd)  — lossless, Slicer-native", "nrrd")
        self.exportFormatCombo.addItem("NIfTI (.nii.gz) — standard neuroimaging", "nii.gz")
        self.exportFormatCombo.addItem("MetaImage (.mha) — ITK standard", "mha")
        imgExportForm.addRow("Format:", self.exportFormatCombo)

        self.exportPetCheck     = qt.QCheckBox("PET volume");           self.exportPetCheck.checked     = True
        self.exportCtCheck      = qt.QCheckBox("CT volume");            self.exportCtCheck.checked      = True
        self.exportInitSegCheck = qt.QCheckBox("Initial segmentation"); self.exportInitSegCheck.checked = True
        self.exportFinalSegCheck= qt.QCheckBox("Final segmentation");   self.exportFinalSegCheck.checked= True
        self.exportMipGifCheck  = qt.QCheckBox("MIP rotating GIF (PET + final segmentation overlay)")
        self.exportMipGifCheck.checked = True
        self.exportMipGifCheck.toolTip = (
            "Generate a 360° rotating Maximum Intensity Projection of the PET "
            "with the final segmentation contours overlaid, saved as an animated GIF."
        )
        imgExportForm.addRow(self.exportPetCheck)
        imgExportForm.addRow(self.exportCtCheck)
        imgExportForm.addRow(self.exportInitSegCheck)
        imgExportForm.addRow(self.exportFinalSegCheck)
        imgExportForm.addRow(self.exportMipGifCheck)

        # MIP GIF options
        mipRow = qt.QHBoxLayout()
        mipFramesLabel = qt.QLabel("Frames:")
        self.mipFramesSpin = qt.QSpinBox()
        self.mipFramesSpin.minimum = 12
        self.mipFramesSpin.maximum = 72
        self.mipFramesSpin.value = 36
        self.mipFramesSpin.toolTip = "Number of rotation frames (36 = every 10°)"
        mipFpsLabel = qt.QLabel("  FPS:")
        self.mipFpsSpin = qt.QSpinBox()
        self.mipFpsSpin.minimum = 4
        self.mipFpsSpin.maximum = 30
        self.mipFpsSpin.value = 12
        mipRow.addWidget(mipFramesLabel)
        mipRow.addWidget(self.mipFramesSpin)
        mipRow.addWidget(mipFpsLabel)
        mipRow.addWidget(self.mipFpsSpin)
        mipRow.addStretch()
        imgExportForm.addRow(mipRow)

        self.exportImagesBtn = qt.QPushButton("  Export Selected Images & Segmentations")
        self.exportImagesBtn.setIcon(qt.QApplication.style().standardIcon(qt.QStyle.SP_DialogSaveButton))
        self.exportImagesBtn.setIconSize(qt.QSize(20, 20))
        self.exportImagesBtn.clicked.connect(self.onExportImages)
        imgExportForm.addRow(self.exportImagesBtn)

        self.exportProgressBar = qt.QProgressBar()
        self.exportProgressBar.setRange(0, 100)
        self.exportProgressBar.value = 0
        self.exportProgressBar.setVisible(False)
        imgExportForm.addRow(self.exportProgressBar)

        self.exportImagesStatusLabel = qt.QLabel("")
        self.exportImagesStatusLabel.setWordWrap(True)
        imgExportForm.addRow(self.exportImagesStatusLabel)

        self.layout.addStretch(1)

        # Internal state
        self._resultsDf = None          # summary DataFrame after run
        self._lesionResultDfs = {}      # per-lesion detail DataFrames
        self._resultSegNode = None      # final multi-label segmentation node

        # Interactive curve state
        self._plotChartNode = None
        self._plotSeriesNode = None
        self._markerSeriesNode = None
        self._currentCurveLesionIdx = None
        self._suppressSliderSignal = False
        self._isDraggingChart = False
        self._chartObserverTags = []
        self._currentCurveValidXY = []   # list of (threshold, volume) for snapping
    # ------------------------------------------------------------------
    # Callbacks
    # ------------------------------------------------------------------

    def onOpenDicomBrowser(self):
        slicer.util.selectModule("DICOM")

    def onOpenSegmentEditor(self):
        slicer.util.selectModule("SegmentEditor")

    def onGenerateOtsu(self):
        petNode = self.petSelector.currentNode()
        if petNode is None:
            slicer.util.warningDisplay("Please select a PET volume first.")
            return
        self.statusLabel.text = "Generating Otsu segmentation…"
        slicer.app.processEvents()
        try:
            segNode = self.logic.generateOtsuSegmentation(
                petNode,
                minObjectVolMl=self.minObjVolSpin.value,
                useSUV=self.useSUVCheckbox.checked,
            )
            self.segNodeSelector.setCurrentNode(segNode)
            self.toggleInitSegBtn.setEnabled(True)
            self.toggleInitSegBtn.text = "👁  Hide Initial Segmentation"
            self._applyViewerLayout()
            self.statusLabel.text = "Otsu segmentation ready. Correct it in Segment Editor, then run."
        except Exception as e:
            slicer.util.errorDisplay(f"Otsu generation failed:\n{e}")
            self.statusLabel.text = "Error during Otsu generation."

    def onRunSegmentation(self):
        petNode = self.petSelector.currentNode()
        segNode = self.segNodeSelector.currentNode()
        if petNode is None:
            slicer.util.warningDisplay("Please select a PET volume.")
            return
        if segNode is None:
            slicer.util.warningDisplay("Please generate (or select) an initial segmentation.")
            return

        self.runBtn.enabled = False
        self.progressBar.value = 0
        self.statusLabel.text = "Running iterative segmentation…"
        slicer.app.processEvents()

        try:
            resultsDf, lesionResultDfs, resultSegNode = self.logic.runIterativeSegmentation(
                petNode=petNode,
                segNode=segNode,
                useSUV=self.useSUVCheckbox.checked,
                plateauThreshold=self.plateauThreshSpin.value,
                postPlateauIterations=self.postPlateauSpin.value,
                iterationRange=self.iterRangeSpin.value,
                lowestVolumeMl=self.lowestVolSpin.value,
                keepLargest=self.keepLargestCheckbox.checked,
                maskToRough=self.maskToRoughCheckbox.checked,
                thresholdStep=self.thresholdStepSpin.value,
                progressCallback=self._updateProgress,
            )
            self._resultsDf = resultsDf
            self._lesionResultDfs = lesionResultDfs
            self._resultSegNode = resultSegNode

            # Apply viewer layout — PET background, CT overlay, PET-DICOM colormap
            self._applyPetColormap(petNode)
            self._applyViewerLayout()

            # Hide the initial segmentation — user can toggle it back
            initSegNode = self.segNodeSelector.currentNode()
            if initSegNode is not None:
                initSegNode.SetDisplayVisibility(0)
                self.toggleInitSegBtn.text = "👁  Show Initial Segmentation"
                self.toggleInitSegBtn.setEnabled(True)

            self.statusLabel.text = f"Done! {len(lesionResultDfs)} lesion(s) segmented."
            self.progressBar.value = 100

            # Show results table immediately without requiring a save
            self._pushResultsToSlicerTable()

            # Populate the interactive curve lesion selector
            self.curveLesionSelector.blockSignals(True)
            self.curveLesionSelector.clear()
            for lesionIdx in sorted(lesionResultDfs.keys()):
                segName = self.logic._lesionCache[lesionIdx].get("segName", f"Lesion {lesionIdx}")
                self.curveLesionSelector.addItem(segName, lesionIdx)
            self.curveLesionSelector.blockSignals(False)
            self.curveStatusLabel.text = "Select a lesion above to inspect/adjust its curve."
            if self.curveLesionSelector.count > 0:
                self.curveLesionSelector.setCurrentIndex(0)
                self.onCurveLesionChanged(0)
        except Exception as e:
            import traceback
            slicer.util.errorDisplay(f"Segmentation failed:\n{e}\n\n{traceback.format_exc()}")
            self.statusLabel.text = "Error during segmentation."
        finally:
            self.runBtn.enabled = True

    def _updateProgress(self, value, message=""):
        self.progressBar.value = int(value)
        if message:
            self.statusLabel.text = message
        slicer.app.processEvents()

    def onPetVolumeSelected(self, node):
        """Immediately apply PET colormap and set viewer layout when PET is chosen."""
        if node is not None:
            self._applyPetColormap(node)
            self._applyViewerLayout()

    def onCtVolumeSelected(self, node):
        """Update viewer overlay when CT is chosen."""
        self._applyViewerLayout()

    def onColormapChanged(self, index):
        """Re-apply the chosen colormap to the current PET volume."""
        petNode = self.petSelector.currentNode()
        if petNode is not None:
            self._applyPetColormap(petNode)

    def _applyPetColormap(self, volumeNode):
        """
        Set the PET volume display to the user-selected colormap and auto window/level.
        """
        if volumeNode is None:
            return
        displayNode = volumeNode.GetScalarVolumeDisplayNode()
        if displayNode is None:
            volumeNode.CreateDefaultDisplayNodes()
            displayNode = volumeNode.GetScalarVolumeDisplayNode()
        if displayNode is None:
            return

        # Get the name chosen in the combo (fallback to PET-DICOM)
        selectedName = self.petColormapCombo.currentData \
            if hasattr(self, "petColormapCombo") else "PET-DICOM"

        # Try the selected name first, then common fallbacks
        colorNode = None
        for name in (selectedName, "PET-DICOM", "PET-Heat", "PET", "Hot", "Grey"):
            colorNode = slicer.mrmlScene.GetFirstNodeByName(name)
            if colorNode is not None:
                break
        if colorNode is not None:
            displayNode.SetAndObserveColorNodeID(colorNode.GetID())

        # Auto window/level across full scalar range
        displayNode.SetAutoWindowLevel(0)
        displayNode.SetAutoThreshold(0)
        if volumeNode.GetImageData() is not None:
            lo, hi = volumeNode.GetImageData().GetScalarRange()
            displayNode.SetWindowLevelMinMax(lo, hi)

    def _applyViewerLayout(self):
        """
        Set all slice viewers so that:
          - Background = PET  (with PET-DICOM colormap, full opacity)
          - Foreground = CT   (greyscale overlay, 30% opacity)  — if CT is loaded
        Resets slice positions and fits all views.
        """
        petNode = self.petSelector.currentNode()
        ctNode  = self.ctSelector.currentNode()

        if petNode is None:
            return  # nothing to show yet

        if ctNode is not None:
            slicer.util.setSliceViewerLayers(
                background=petNode,
                foreground=ctNode,
                foregroundOpacity=0.3,
            )
        else:
            slicer.util.setSliceViewerLayers(
                background=petNode,
                foreground=None,
                foregroundOpacity=0.0,
            )

        slicer.util.resetSliceViews()

    def onToggleInitialSegmentation(self):
        initSegNode = self.segNodeSelector.currentNode()
        if initSegNode is None:
            return
        currentlyVisible = initSegNode.GetDisplayVisibility()
        initSegNode.SetDisplayVisibility(0 if currentlyVisible else 1)
        self.toggleInitSegBtn.text = (
            "👁  Hide Initial Segmentation" if not currentlyVisible
            else "👁  Show Initial Segmentation"
        )

    # ------------------------------------------------------------------
    # Interactive plateau curve callbacks
    # ------------------------------------------------------------------

    def _installChartMouseObservers(self):
        """
        Attach LeftButtonPress / MouseMove / LeftButtonRelease observers to
        the chart's render window interactor so the user can click or drag
        anywhere on the chart to set the threshold directly.
        """
        try:
            iren = self.chartView.renderWindow().GetInteractor()
        except Exception:
            return

        # Remove any previously installed observers
        for tag in self._chartObserverTags:
            try:
                iren.RemoveObserver(tag)
            except Exception:
                pass
        self._chartObserverTags = []

        t1 = iren.AddObserver("LeftButtonPressEvent",   self._onChartMousePress)
        t2 = iren.AddObserver("MouseMoveEvent",         self._onChartMouseMove)
        t3 = iren.AddObserver("LeftButtonReleaseEvent", self._onChartMouseRelease)
        self._chartObserverTags = [t1, t2, t3]

    def _chartPixelToThreshold(self, iren):
        """
        Convert the current interactor pixel position to a threshold value by
        mapping through the chart's axis transform, then snapping to the
        nearest recorded threshold in the current lesion's curve.
        Returns the snapped threshold float, or None if not possible.
        """
        if not self._currentCurveValidXY:
            return None
        try:
            chart = self.chartView.chart()
            xAxis = chart.GetAxis(1)   # bottom = X axis
            yAxis = chart.GetAxis(0)   # left   = Y axis

            # Screen pixel position of the mouse
            px, py = iren.GetEventPosition()

            # Chart geometry in pixels (set by VTK during paint)
            # GetPoint1/Point2 give the plot area corners in scene coords
            p1x = chart.GetPoint1()[0]
            p2x = chart.GetPoint2()[0]
            p1y = chart.GetPoint1()[1]
            p2y = chart.GetPoint2()[1]

            if p2x <= p1x or p2y <= p1y:
                return None

            # Map pixel → data coordinate linearly
            xMin, xMax = xAxis.GetMinimum(), xAxis.GetMaximum()
            dataX = xMin + (px - p1x) / (p2x - p1x) * (xMax - xMin)

            # Snap to nearest recorded threshold on the curve
            nearest = min(self._currentCurveValidXY,
                          key=lambda tv: abs(tv[0] - dataX))
            return float(nearest[0])
        except Exception:
            return None

    def _onChartMousePress(self, obj, event):
        thresh = self._chartPixelToThreshold(obj)
        if thresh is not None:
            self._isDraggingChart = True
            self._suppressSliderSignal = True
            self.thresholdSlider.value = thresh
            self._suppressSliderSignal = False
            self.onThresholdSliderChanged(thresh)

    def _onChartMouseMove(self, obj, event):
        if not self._isDraggingChart:
            return
        thresh = self._chartPixelToThreshold(obj)
        if thresh is not None:
            self._suppressSliderSignal = True
            self.thresholdSlider.value = thresh
            self._suppressSliderSignal = False
            self.onThresholdSliderChanged(thresh)

    def _onChartMouseRelease(self, obj, event):
        self._isDraggingChart = False

    def onCurveLesionChanged(self, index):
        if index < 0:
            return
        lesionIdx = self.curveLesionSelector.itemData(index)
        if lesionIdx is None:
            return
        self._currentCurveLesionIdx = lesionIdx
        self._buildPlotForLesion(lesionIdx)

    def _buildPlotForLesion(self, lesionIdx):
        try:
            thresholds, volumes, unit, autoThreshold = self.logic.getLesionCurve(lesionIdx)
        except Exception as e:
            self.curveStatusLabel.text = f"Could not load curve: {e}"
            return

        import vtk as vtkmod

        cache = self.logic._lesionCache[lesionIdx]
        segName = cache.get("segName", f"Lesion {lesionIdx}")
        currentThresh = cache["currentThreshold"]
        currentVol    = cache["currentStats"]["volume_mL"]

        chart = self.chartView.chart()
        chart.ClearPlots()
        chart.SetTitle(f"{segName}  —  Volume vs. Threshold")
        chart.GetAxis(0).SetTitle("Volume (mL)")
        chart.GetAxis(1).SetTitle(f"Threshold ({unit})")
        chart.SetShowLegend(True)

        # ── full convergence table ─────────────────────────────────────
        table = vtkmod.vtkTable()
        tCol = vtkmod.vtkFloatArray(); tCol.SetName("Threshold")
        vCol = vtkmod.vtkFloatArray(); vCol.SetName("Volume_mL")
        valid = [(float(t), float(v)) for t, v in zip(thresholds, volumes)
                 if t == t and v == v]
        for t, v in valid:
            tCol.InsertNextValue(t)
            vCol.InsertNextValue(v)
        table.AddColumn(tCol)
        table.AddColumn(vCol)

        # ── convergence line ──────────────────────────────────────────
        line = chart.AddPlot(0)           # vtkChart::LINE
        line.SetInputData(table, 0, 1)
        line.SetLabel(f"{segName} — volume curve")
        line.SetColor(51, 128, 230, 255)
        line.SetWidth(2.0)

        # ── helper: make a visible scatter point ──────────────────────
        def _makeMarker(thresh, vol, r, g, b, label, size=16, style=None,
                        penColor=None, penWidth=2.0):
            t = vtkmod.vtkTable()
            cx = vtkmod.vtkFloatArray(); cx.SetName("X")
            cy = vtkmod.vtkFloatArray(); cy.SetName("Y")
            # Two identical rows — VTK skips single-row scatter in UpdateCache
            cx.InsertNextValue(float(thresh)); cy.InsertNextValue(float(vol))
            cx.InsertNextValue(float(thresh)); cy.InsertNextValue(float(vol))
            t.AddColumn(cx); t.AddColumn(cy)
            p = chart.AddPlot(3)           # vtkChart::POINTS
            p.SetInputData(t, 0, 1)
            p.SetLabel(label)
            p.GetBrush().SetColor(r, g, b, 255)
            # Pen color matches fill so it doesn't override the visible color
            pr, pg, pb = penColor if penColor is not None else (r, g, b)
            p.GetPen().SetColor(pr, pg, pb, 255)
            p.GetPen().SetWidth(penWidth)
            # Explicitly bind axes — required for scatter plots to render
            p.SetXAxis(chart.GetAxis(1))   # bottom
            p.SetYAxis(chart.GetAxis(0))   # left
            pp = vtkmod.vtkPlotPoints.SafeDownCast(p)
            if pp is not None:
                markerStyle = style if style is not None else vtkmod.vtkPlotPoints.CIRCLE
                pp.SetMarkerStyle(markerStyle)
                pp.SetMarkerSize(float(size))
            return p, t

        # ── auto-plateau marker — large green cross with white halo ───
        closest = min(valid, key=lambda tv: abs(tv[0] - autoThreshold))
        autoVol = closest[1]

        # White halo (larger, behind)
        _makeMarker(autoThreshold, autoVol, 255, 255, 255, "_auto_halo",
                    size=48, style=vtkmod.vtkPlotPoints.CIRCLE,
                    penColor=(255, 255, 255))
        # Bright green cross (on top) — 2× line width via pen
        _, autoTable = _makeMarker(
            autoThreshold, autoVol, 30, 180, 60,
            f"★ Auto plateau  {autoThreshold:.4f}",
            size=36, style=vtkmod.vtkPlotPoints.CROSS,
            penColor=(30, 180, 60), penWidth=4.0,
        )

        # ── selected-threshold marker (dark red circle, draggable) ────
        redScatter, markerTable = _makeMarker(
            currentThresh, currentVol, 160, 20, 20,
            f"● Selected  {currentThresh:.4f}",
            size=24, style=vtkmod.vtkPlotPoints.CIRCLE,
            penColor=(160, 20, 20), penWidth=4.0,
        )

        # Force VTK context scene to repaint with all new plots
        chart.RecalculateBounds()
        scene = chart.GetScene()
        if scene is not None:
            scene.SetDirty(True)
        self.chartView.renderWindow().Render()
        self.chartView.repaint()

        # Store valid curve XY pairs so the click handler can snap to them
        self._currentCurveValidXY = valid

        # Store refs for live updates
        self._currentVTKMarkerTable = markerTable
        self._currentVTKChart = chart
        self._currentVTKMarkerScatter = redScatter

        # ── slider ────────────────────────────────────────────────────
        validThresholds = [t for t, _ in valid]
        tMin = float(min(validThresholds))
        tMax = float(max(validThresholds))

        # Extend the max to the actual PET peak in the ROI so the user can
        # explore thresholds above the recorded iteration range freely.
        petMax = float(np.max(cache["petCropArr"][cache["roughMaskArr"] > 0])) \
            if np.any(cache["roughMaskArr"] > 0) else tMax
        sliderMax = max(tMax, petMax)

        step = max((sliderMax - tMin) / 300.0, 0.0001)
        self._suppressSliderSignal = True
        self.thresholdSlider.minimum = tMin
        self.thresholdSlider.maximum = sliderMax if sliderMax > tMin else tMin + step
        self.thresholdSlider.singleStep = step
        self.thresholdSlider.value = currentThresh
        self._suppressSliderSignal = False
        self.thresholdSlider.setEnabled(True)
        self.applyThresholdBtn.setEnabled(True)

        self._updateCurveStatsLabels(cache["currentStats"])
        nPoints = len(valid)
        self.curveStatusLabel.text = (
            f"{segName}: {nPoints} iterations plotted (full range). "
            f"★ Auto plateau at {autoThreshold:.4f} {unit} (green cross). "
            f"Drag slider to move ● red marker and override."
        )

    def onThresholdSliderChanged(self, value):
        if self._suppressSliderSignal or self._currentCurveLesionIdx is None:
            return
        lesionIdx = self._currentCurveLesionIdx
        try:
            stats = self.logic.rethresholdLesion(lesionIdx, value)
            self.logic.updateLivePreview(lesionIdx)
            self._updateCurveStatsLabels(stats)
            self._updateMarkerPosition(lesionIdx, value, stats["volume_mL"])
        except Exception as e:
            self.curveStatusLabel.text = f"Re-threshold failed: {e}"

    def _updateMarkerPosition(self, lesionIdx, threshold, volume):
        if not hasattr(self, "_currentVTKMarkerTable") or self._currentVTKMarkerTable is None:
            return
        table = self._currentVTKMarkerTable
        if table.GetNumberOfRows() == 0:
            return
        for row in range(table.GetNumberOfRows()):
            table.GetColumn(0).SetTuple1(row, float(threshold))
            table.GetColumn(1).SetTuple1(row, float(volume))
        table.Modified()
        # Update the legend label to show the new threshold value
        if hasattr(self, "_currentVTKMarkerScatter") and self._currentVTKMarkerScatter is not None:
            self._currentVTKMarkerScatter.SetLabel(f"● Selected  {threshold:.4f}")
        # Mark scene dirty so VTK repaints
        chart = self.chartView.chart()
        chart.RecalculateBounds()
        scene = chart.GetScene()
        if scene is not None:
            scene.SetDirty(True)
        self.chartView.renderWindow().Render()
        self.chartView.repaint()

    def _updateCurveStatsLabels(self, stats):
        self.curveVolumeLabel.text = f"Volume: {stats['volume_mL']:.3f} mL"
        self.curveMeanLabel.text = f"Mean: {stats['mean']:.4f}"

    def onResetToAutoPlateau(self):
        if self._currentCurveLesionIdx is None:
            return
        lesionIdx = self._currentCurveLesionIdx
        cache = self.logic._lesionCache.get(lesionIdx)
        if cache is None:
            return
        self.thresholdSlider.value = cache["autoThreshold"]  # triggers onThresholdSliderChanged

    def onApplyThreshold(self):
        if self._currentCurveLesionIdx is None or self._resultsDf is None:
            return
        lesionIdx = self._currentCurveLesionIdx
        try:
            self._resultsDf = self.logic.commitLesionThreshold(lesionIdx, self._resultsDf)
            # Mark that this lesion was manually overridden — used in convergence plot
            self.logic._lesionCache[lesionIdx]["manuallyOverridden"] = True
            self.curveStatusLabel.text = (
                f"Applied. Lesion {lesionIdx} final threshold updated in the results table."
            )
            # Refresh the live table view
            self._pushResultsToSlicerTable()
        except Exception as e:
            slicer.util.errorDisplay(f"Failed to apply threshold:\n{e}")

    def onShowResultsTable(self):
        if self._resultsDf is None:
            slicer.util.warningDisplay("No results yet. Run segmentation first.")
            return
        self._pushResultsToSlicerTable()

    def _pushResultsToSlicerTable(self):
        """Push the current summary DataFrame into a vtkMRMLTableNode and show it."""
        if self._resultsDf is None:
            return
        import vtk as vtkmod

        # Reuse the same table node across refreshes
        tableNode = slicer.mrmlScene.GetFirstNodeByName("__xeos_results_table__")
        if tableNode is None:
            tableNode = slicer.mrmlScene.AddNewNodeByClass(
                "vtkMRMLTableNode", "__xeos_results_table__"
            )
        tableNode.RemoveAllColumns()

        table = tableNode.GetTable()
        table.Initialize()

        for col in self._resultsDf.columns:
            arr = vtkmod.vtkStringArray()
            arr.SetName(str(col))
            for val in self._resultsDf[col]:
                arr.InsertNextValue(str(val) if val is not None and val == val else "")
            table.AddColumn(arr)

        # Switch to the Tables module and select the node so user sees it immediately
        slicer.app.layoutManager().setLayout(
            slicer.modules.tables.logic().GetLayoutWithTable(
                slicer.app.layoutManager().layout
            )
        )
        slicer.app.applicationLogic().GetSelectionNode().SetActiveTableID(tableNode.GetID())
        slicer.app.applicationLogic().PropagateTableSelection()

    def onExportExcel(self):
        if self._resultsDf is None:
            slicer.util.warningDisplay("No results to export. Please run segmentation first.")
            return
        outputDir = self.outputDirBtn.directory
        if not outputDir:
            slicer.util.warningDisplay("Please select an output folder.")
            return
        try:
            paths = self.logic.exportToExcel(
                resultsDf=self._resultsDf,
                lesionResultDfs=self._lesionResultDfs,
                outputDir=outputDir,
            )
            # Optionally save convergence plot PNG alongside the Excel files
            if self.saveConvergencePlotCheck.checked:
                try:
                    plotPath = self.logic.saveConvergencePlot(outputDir)
                    paths.append(plotPath)
                except Exception as pe:
                    slicer.util.warningDisplay(f"Excel saved OK, but convergence plot failed:\n{pe}")

            slicer.util.infoDisplay(
                "Saved:\n" + "\n".join(paths),
                windowTitle="Export complete",
            )
        except Exception as e:
            slicer.util.errorDisplay(f"Export failed:\n{e}")

    def onExportImages(self):
        outputDir = self.outputDirBtn.directory   # shared folder
        if not outputDir:
            slicer.util.warningDisplay("Please select an output folder in Step 6 first.")
            return

        fmt = self.exportFormatCombo.currentData

        exportItems = []
        if self.exportPetCheck.checked:
            node = self.petSelector.currentNode()
            if node:
                exportItems.append(("PET", node, False))
            else:
                slicer.util.warningDisplay("No PET volume selected — skipping PET export.")
        if self.exportCtCheck.checked:
            node = self.ctSelector.currentNode()
            if node:
                exportItems.append(("CT", node, False))
            else:
                slicer.util.warningDisplay("No CT volume selected — skipping CT export.")
        if self.exportInitSegCheck.checked:
            node = self.segNodeSelector.currentNode()
            if node:
                exportItems.append(("InitialSegmentation", node, True))
            else:
                slicer.util.warningDisplay("No initial segmentation selected — skipping.")
        if self.exportFinalSegCheck.checked:
            node = self._resultSegNode
            if node:
                exportItems.append(("FinalSegmentation", node, True))
            else:
                slicer.util.warningDisplay("No final segmentation — run iterative segmentation first.")

        doGif = (self.exportMipGifCheck.checked
                 and self.petSelector.currentNode() is not None
                 and self._resultSegNode is not None)

        if not exportItems and not doGif:
            slicer.util.warningDisplay("Nothing to export. Check your selections.")
            return

        # Calculate total steps for progress bar
        totalSteps = len(exportItems) + (self.mipFramesSpin.value * 2 if doGif else 0)

        self.exportProgressBar.setVisible(True)
        self.exportProgressBar.value = 0
        self.exportImagesBtn.setEnabled(False)
        savedPaths = []
        step = 0

        def _progress(pct, msg=""):
            self.exportProgressBar.value = int(pct)
            if msg:
                self.exportImagesStatusLabel.text = msg
            slicer.app.processEvents()

        try:
            # Export volumes / segmentations one by one with individual updates
            for i, (label, node, isSeg) in enumerate(exportItems):
                _progress(
                    int(i / totalSteps * 100),
                    f"Saving {label}…"
                )
                paths = self.logic.exportImages(
                    exportItems=[(label, node, isSeg)],
                    outputDir=outputDir,
                    fmt=fmt,
                )
                savedPaths.extend(paths)
                step += 1

            # MIP GIF with frame-level progress
            if doGif:
                nFrames = self.mipFramesSpin.value
                baseStep = step / totalSteps * 100

                def gifProgress(frameDone, totalFrames, msg=""):
                    pct = baseStep + (frameDone / totalFrames) * (100 - baseStep)
                    _progress(pct, msg or f"MIP GIF: frame {frameDone}/{totalFrames}…")

                _progress(baseStep, "Generating MIP GIF…")
                gifPath = self.logic.generateMipGif(
                    petNode=self.petSelector.currentNode(),
                    segNode=self._resultSegNode,
                    outputDir=outputDir,
                    nFrames=nFrames,
                    fps=self.mipFpsSpin.value,
                    progressCallback=gifProgress,
                )
                savedPaths.append(gifPath)

            _progress(100, f"Done — {len(savedPaths)} file(s) saved.")
            slicer.util.infoDisplay(
                "Exported:\n" + "\n".join(savedPaths),
                windowTitle="Image export complete",
            )
        except Exception as e:
            import traceback
            slicer.util.errorDisplay(f"Image export failed:\n{e}\n\n{traceback.format_exc()}")
            self.exportImagesStatusLabel.text = "Export failed — see error dialog."
        finally:
            self.exportImagesBtn.setEnabled(True)
            self.exportProgressBar.setVisible(False)


# ─────────────────────────────────────────────────────────────────────────────
# Logic
# ─────────────────────────────────────────────────────────────────────────────
class XEOSIterativeSegmentationLogic(ScriptedLoadableModuleLogic):
    """All computation — no Qt / GUI here."""

    def __init__(self):
        super().__init__()
        # Cache populated by runIterativeSegmentation(), keyed by lesionIdx.
        # Enables instant re-thresholding for the interactive plateau curve
        # without re-cropping or re-running the full iterative search.
        self._lesionCache = {}
        # Geometry needed to resample a cropped result back into full PET space
        self._petSitkRef = None
        self._resultSegNode = None
        self._resultLabelArr = None
        self._useSUV = False

    # ------------------------------------------------------------------
    # Helpers: Slicer volume ↔ numpy
    # ------------------------------------------------------------------

    @staticmethod
    def _volumeToArray(volumeNode):
        """Return (array_zyx, spacing_xyz, origin_xyz, direction_matrix)."""
        import sitkUtils
        sitkImage = sitkUtils.PullVolumeFromSlicer(volumeNode)
        arr = sitk_GetArrayFromImage(sitkImage)
        return arr, sitkImage.GetSpacing(), sitkImage.GetOrigin(), sitkImage.GetDirection()

    @staticmethod
    def _arrayToVolume(arr, referenceNode, name):
        """Create a new vtkMRMLScalarVolumeNode from a numpy array."""
        import sitkUtils
        import SimpleITK as sitk
        refSitk = sitkUtils.PullVolumeFromSlicer(referenceNode)
        newSitk = sitk.GetImageFromArray(arr.astype(np.float32))
        newSitk.CopyInformation(refSitk)
        newNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLScalarVolumeNode", name)
        sitkUtils.PushVolumeToSlicer(newSitk, newNode)
        return newNode

    # ------------------------------------------------------------------
    # Otsu initial segmentation
    # ------------------------------------------------------------------

    def generateOtsuSegmentation(self, petNode, minObjectVolMl=0.1, useSUV=False):
        """
        1. Pull PET as SimpleITK image
        2. Otsu threshold
        3. Remove small objects
        4. Binary dilate (3 voxels)
        5. Connected-component label
        6. Push result into a vtkMRMLSegmentationNode (one segment per label)
        """
        import SimpleITK as sitk
        import sitkUtils

        petSitk = sitkUtils.PullVolumeFromSlicer(petNode)

        if not useSUV:
            # Convert Bq/mL → kBq/mL for consistent thresholding
            petSitk = sitk.Cast(petSitk, sitk.sitkFloat64) * 0.001

        petSitk = sitk.Cast(petSitk, sitk.sitkFloat32)

        # Otsu threshold
        otsuFilter = sitk.OtsuThresholdImageFilter()
        otsuFilter.SetInsideValue(0)
        otsuFilter.SetOutsideValue(1)
        binaryMask = otsuFilter.Execute(petSitk)
        binaryMask = sitk.Cast(binaryMask, sitk.sitkUInt8)

        # Remove objects below minimum volume
        spacing = petSitk.GetSpacing()
        voxelVolMl = spacing[0] * spacing[1] * spacing[2] / 1000.0
        minVoxels = int(minObjectVolMl / voxelVolMl) if voxelVolMl > 0 else 0

        if minVoxels > 0:
            binaryMask = sitk.RelabelComponent(
                sitk.ConnectedComponent(binaryMask), minimumObjectSize=minVoxels
            )
            binaryMask = sitk.Cast(binaryMask > 0, sitk.sitkUInt8)

        # Dilate slightly
        binaryMask = sitk.BinaryDilate(binaryMask, [3, 3, 3])

        # Label connected components
        labeled = sitk.ConnectedComponent(binaryMask)
        labeled = sitk.RelabelComponent(labeled, sortByObjectSize=True)
        labeled = sitk.Cast(labeled, sitk.sitkUInt16)

        # Create a label map volume node
        labelNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLLabelMapVolumeNode",
                                                        "OtsuInitialSegmentation")
        sitkUtils.PushVolumeToSlicer(labeled, labelNode)

        # Convert label map → segmentation node
        segNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentationNode",
                                                      "InitialSegmentation")
        slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(
            labelNode, segNode
        )
        segNode.CreateClosedSurfaceRepresentation()

        # Remove temporary label node
        slicer.mrmlScene.RemoveNode(labelNode)

        return segNode

    # ------------------------------------------------------------------
    # Core iterative segmentation
    # ------------------------------------------------------------------

    def runIterativeSegmentation(
        self,
        petNode,
        segNode,
        useSUV=False,
        plateauThreshold=2.0,
        postPlateauIterations=10,
        iterationRange=100,
        lowestVolumeMl=0.0,
        keepLargest=True,
        maskToRough=True,
        thresholdStep=0.1,
        progressCallback=None,
    ):
        """
        For each label in the segmentation node, run iterative thresholding
        and detect the volume plateau.

        Returns
        -------
        resultsDf : pd.DataFrame  — one row per lesion with final metrics
        lesionResultDfs : dict    — per-lesion iteration DataFrames
        resultSegNode : vtkMRMLSegmentationNode — final multi-label result
        """
        import SimpleITK as sitk
        import sitkUtils
        import pandas as pd

        def _progress(pct, msg=""):
            if progressCallback:
                progressCallback(pct, msg)

        # Pull PET
        petSitk = sitkUtils.PullVolumeFromSlicer(petNode)
        petSitk = sitk.Cast(petSitk, sitk.sitkFloat64)
        if not useSUV:
            petSitk = petSitk * 0.001  # Bq/mL → kBq/mL

        # Export segmentation to a label map that matches PET geometry
        labelNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLLabelMapVolumeNode",
                                                        "__xeos_tmp_label__")
        slicer.modules.segmentations.logic().ExportVisibleSegmentsToLabelmapNode(
            segNode, labelNode, petNode
        )
        labelSitk = sitkUtils.PullVolumeFromSlicer(labelNode)
        slicer.mrmlScene.RemoveNode(labelNode)

        labelSitk = sitk.Cast(labelSitk, sitk.sitkUInt16)
        labelArr = sitk.GetArrayFromImage(labelSitk)

        lesionIndices = [int(v) for v in np.unique(labelArr) if v != 0]
        if not lesionIndices:
            raise ValueError("No lesion labels found in the segmentation.")

        resultsDf = pd.DataFrame()
        lesionResultDfs = {}
        self._lesionCache = {}

        # Result label image (same space as PET)
        resultLabelArr = np.zeros_like(labelArr, dtype=np.uint16)

        for li, lesionIdx in enumerate(lesionIndices):
            _progress(
                int(li / len(lesionIndices) * 90),
                f"Lesion {lesionIdx} ({li+1}/{len(lesionIndices)}) — step {thresholdStep}…"
            )

            # Isolate this lesion
            lesionMaskArr = (labelArr == lesionIdx).astype(np.uint8)
            lesionMaskSitk = sitk.GetImageFromArray(lesionMaskArr)
            lesionMaskSitk.CopyInformation(labelSitk)

            # Crop PET + mask to bounding box
            petCropped, maskCropped = self._cropToBoundingBox(petSitk, lesionMaskSitk)

            # Iterative threshold search
            iterDf, finalSegSitk, finalThresh, finalStats = self._searchForPlateau(
                petCropped=petCropped,
                roughMaskCropped=maskCropped,
                lesionIndex=lesionIdx,
                useSUV=useSUV,
                plateauThreshold=plateauThreshold,
                postPlateauIterations=postPlateauIterations,
                iterationRange=iterationRange,
                lowestVolumeMl=lowestVolumeMl,
                maskToRough=maskToRough,
                thresholdStep=thresholdStep,
            )

            lesionResultDfs[lesionIdx] = iterDf

            # Keep largest object
            if keepLargest:
                finalSegSitk = self._keepLargestComponent(finalSegSitk)

            # Resample result back to full PET space
            finalSegFullSitk = sitk.Resample(
                sitk.Cast(finalSegSitk, sitk.sitkUInt8),
                petSitk,
                sitk.Transform(),
                sitk.sitkNearestNeighbor,
                0,
                sitk.sitkUInt8,
            )
            finalArr = sitk.GetArrayFromImage(finalSegFullSitk)
            resultLabelArr[finalArr > 0] = lesionIdx

            # Pull the human-readable segment name.
            # Use the label map we already exported — ask the segmentation logic
            # for the label→name mapping via GetSegmentIDs and the shared labelling.
            segName = f"Lesion {lesionIdx}"  # guaranteed fallback
            try:
                segmentation = segNode.GetSegmentation()
                # GetSegmentIDs returns all segment IDs in order; the label map
                # assigns values 1,2,3... in that same order when exported via
                # ExportVisibleSegmentsToLabelmapNode — so label value == position+1.
                segIds = vtk.vtkStringArray()
                segmentation.GetSegmentIDs(segIds)
                for si in range(segIds.GetNumberOfValues()):
                    if si + 1 == lesionIdx:
                        sid = segIds.GetValue(si)
                        segName = segmentation.GetSegment(sid).GetName()
                        break
            except Exception:
                pass  # keep fallback name

            # Collect summary metrics — segment_name first so it's the leftmost column
            row = {
                "segment_name": segName,
                "lesion_index": lesionIdx,
                "final_threshold_SUV" if useSUV else "final_threshold_kBqmL": round(float(finalThresh), 4),
                "volume_mL": round(float(finalStats["volume_mL"]), 3),
                "mean_SUV" if useSUV else "mean_kBqmL": round(float(finalStats["mean"]), 4),
                "max_SUV"  if useSUV else "max_kBqmL":  round(float(finalStats["max"]),  4),
                "min_SUV"  if useSUV else "min_kBqmL":  round(float(finalStats["min"]),  4),
                "median_SUV" if useSUV else "median_kBqmL": round(float(finalStats["median"]), 4),
                "std_SUV"  if useSUV else "std_kBqmL":  round(float(finalStats["std"]),  4),
            }
            resultsDf = pd.concat([resultsDf, pd.DataFrame([row])], ignore_index=True)

            # Cache everything needed for instant interactive re-thresholding
            petCropArr = sitk.GetArrayFromImage(petCropped).astype(np.float64)
            roughMaskArr = sitk.GetArrayFromImage(maskCropped).astype(np.uint8)
            spacing = petCropped.GetSpacing()
            voxelVolMl = spacing[0] * spacing[1] * spacing[2] / 1000.0

            self._lesionCache[lesionIdx] = dict(
                petCropped=petCropped,
                petCropArr=petCropArr,
                roughMaskArr=roughMaskArr,
                voxelVolMl=voxelVolMl,
                iterDf=iterDf,
                autoThreshold=float(finalThresh),
                autoStats=finalStats,
                currentThreshold=float(finalThresh),
                currentStats=finalStats,
                maskToRough=maskToRough,
                keepLargest=keepLargest,
                segName=segName,
            )

        # Push final multi-label result as a Segmentation node
        resultLabelSitk = sitk.GetImageFromArray(resultLabelArr)
        resultLabelSitk.CopyInformation(petSitk)
        resultLabelSitk = sitk.Cast(resultLabelSitk, sitk.sitkUInt16)

        import sitkUtils
        tmpLabelNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLLabelMapVolumeNode",
                                                           "__xeos_result_tmp__")
        sitkUtils.PushVolumeToSlicer(resultLabelSitk, tmpLabelNode)

        resultSegNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSegmentationNode",
                                                            "IterativeSegmentation_Result")
        slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(
            tmpLabelNode, resultSegNode
        )
        resultSegNode.CreateClosedSurfaceRepresentation()
        slicer.mrmlScene.RemoveNode(tmpLabelNode)

        # Keep references so the interactive curve can live-update this exact
        # segmentation node/label array without rebuilding everything.
        self._petSitkRef = petSitk
        self._resultSegNode = resultSegNode
        self._resultLabelArr = resultLabelArr
        self._useSUV = useSUV

        _progress(100, "Iterative segmentation complete.")
        return resultsDf, lesionResultDfs, resultSegNode

    # ------------------------------------------------------------------
    # Plateau search
    # ------------------------------------------------------------------

    def _searchForPlateau(
        self,
        petCropped,
        roughMaskCropped,
        lesionIndex,
        useSUV,
        plateauThreshold,
        postPlateauIterations,
        iterationRange,
        lowestVolumeMl,
        maskToRough,
        thresholdStep=0.1,
    ):
        import SimpleITK as sitk
        import pandas as pd

        petArr = sitk.GetArrayFromImage(petCropped).astype(np.float64)
        maskArr = sitk.GetArrayFromImage(roughMaskCropped).astype(np.uint8)
        spacing = petCropped.GetSpacing()
        voxelVolMl = spacing[0] * spacing[1] * spacing[2] / 1000.0

        def _volume_ml(binaryArr):
            return float(np.sum(binaryArr > 0)) * voxelVolMl

        def _stats(binaryArr):
            vals = petArr[binaryArr > 0]
            if len(vals) == 0:
                return dict(mean=0, min=0, max=0, median=0, std=0, volume_mL=0)
            return dict(
                mean=float(np.mean(vals)),
                min=float(np.min(vals)),
                max=float(np.max(vals)),
                median=float(np.median(vals)),
                std=float(np.std(vals)),
                volume_mL=float(len(vals)) * voxelVolMl,
            )

        initialStats = _stats(maskArr)
        roiVoxels = petArr[maskArr > 0]
        maxVal = float(np.max(roiVoxels)) if len(roiVoxels) > 0 else 1.0
        # Start from the actual minimum value inside the ROI — no wasted
        # iterations below the signal floor
        minVal = float(np.min(roiVoxels)) if len(roiVoxels) > 0 else 0.0

        unit = "SUV" if useSUV else "kBq/mL"
        lesionIdx = lesionIndex
        records = []
        records.append({
            "lesion_index": lesionIdx,
            "iteration": 0,
            f"lower_threshold_{unit}": round(minVal, 4),
            "volume_mL": initialStats["volume_mL"],
            f"mean_{unit}": initialStats["mean"],
            "volume_change_pct": None,
            "comment": (
                f"Initial (Otsu/manual) segment. "
                f"ROI min: {minVal:.3f}, max: {maxVal:.3f}. "
                f"Step: {thresholdStep}"
            ),
        })

        # Build threshold array: start at ROI min, step by user-defined size,
        # stop at ROI max, cap at iterationRange steps
        thresholds = np.arange(
            minVal + thresholdStep,
            maxVal,
            thresholdStep,
        )[:iterationRange]

        plateau_reached = False
        post_plateau_counter = 0
        previous_volume = None
        last_seg_arr = maskArr.copy()
        last_thresh = minVal
        last_stats = initialStats

        # These capture the auto-detected plateau result without stopping the loop
        auto_thresh = None
        auto_stats  = None
        auto_seg_arr = None

        for iteration, thresh in enumerate(thresholds, start=1):
            newSeg = (petArr > thresh).astype(np.uint8)
            if maskToRough:
                newSeg[maskArr == 0] = 0

            volume = _volume_ml(newSeg)
            st     = _stats(newSeg)
            comment = ""

            # Minimum-volume floor — lock in the LAST VALID segmentation
            # (the one just before volume dropped below the floor) and keep
            # iterating so the full curve is still visible in the plot.
            if lowestVolumeMl > 0 and volume <= lowestVolumeMl:
                comment = f"Below minimum volume {lowestVolumeMl} mL ← auto selected here"
                if auto_thresh is None:
                    # Use the previous iteration's result, not this one
                    auto_thresh  = last_thresh
                    auto_stats   = last_stats
                    auto_seg_arr = last_seg_arr.copy()

            if previous_volume is not None and previous_volume > 0:
                vol_change_pct = abs((volume - previous_volume) / previous_volume) * 100.0
            else:
                vol_change_pct = None

            # Plateau detection — only applies if minimum volume not yet hit
            if auto_thresh is None:
                if vol_change_pct is not None and vol_change_pct < plateauThreshold:
                    if not plateau_reached:
                        plateau_reached = True
                        comment = f"Plateau detected (Δvol {vol_change_pct:.2f}%)"
                    else:
                        post_plateau_counter += 1
                        comment = f"Post-plateau #{post_plateau_counter}"

                    if post_plateau_counter == postPlateauIterations:
                        auto_thresh  = float(thresh)
                        auto_stats   = st
                        auto_seg_arr = newSeg.copy()
                        comment += " ← auto selected"
                else:
                    plateau_reached      = False
                    post_plateau_counter = 0

            records.append({
                "lesion_index": lesionIdx,
                "iteration": iteration,
                f"lower_threshold_{unit}": round(float(thresh), 6),
                "volume_mL": volume,
                f"mean_{unit}": st["mean"],
                "volume_change_pct": vol_change_pct,
                "comment": comment,
            })

            previous_volume = volume
            last_seg_arr    = newSeg
            last_thresh     = float(thresh)
            last_stats      = st

        # If plateau was never reached, use the last recorded point
        if auto_thresh is None:
            auto_thresh  = last_thresh
            auto_stats   = last_stats
            auto_seg_arr = last_seg_arr

        iterDf = pd.DataFrame(records)

        # The segmentation result is the auto-detected plateau point —
        # the full iterDf still contains all iterations for the plot.
        finalSegSitk = sitk.GetImageFromArray(auto_seg_arr.astype(np.uint8))
        finalSegSitk.CopyInformation(petCropped)

        return iterDf, finalSegSitk, auto_thresh, auto_stats

    # ------------------------------------------------------------------
    # Interactive plateau curve support
    # ------------------------------------------------------------------

    def getLesionCurve(self, lesionIdx):
        """
        Returns (thresholds, volumes, unit_label, autoThreshold) for plotting.
        thresholds/volumes are numpy arrays from the recorded iteration table.
        """
        cache = self._lesionCache.get(lesionIdx)
        if cache is None:
            raise ValueError(f"No cached data for lesion {lesionIdx}. Run segmentation first.")
        iterDf = cache["iterDf"]
        unit = "SUV" if self._useSUV else "kBq/mL"
        threshCol = f"lower_threshold_{unit}"
        thresholds = iterDf[threshCol].to_numpy(dtype=float)
        volumes = iterDf["volume_mL"].to_numpy(dtype=float)
        return thresholds, volumes, unit, cache["autoThreshold"]

    def rethresholdLesion(self, lesionIdx, threshold):
        """
        Fast, in-memory re-threshold of a single lesion using cached cropped
        arrays. Does NOT touch the Slicer scene — call updateLivePreview()
        (or commitLesionThreshold()) to push the change into the viewer.

        Returns dict of stats: mean, min, max, median, std, volume_mL
        """
        import SimpleITK as sitk

        cache = self._lesionCache.get(lesionIdx)
        if cache is None:
            raise ValueError(f"No cached data for lesion {lesionIdx}. Run segmentation first.")

        petArr = cache["petCropArr"]
        roughMaskArr = cache["roughMaskArr"]
        voxelVolMl = cache["voxelVolMl"]

        newSeg = (petArr > threshold).astype(np.uint8)
        if cache["maskToRough"]:
            newSeg[roughMaskArr == 0] = 0

        if cache["keepLargest"] and np.any(newSeg):
            segSitk = sitk.GetImageFromArray(newSeg)
            segSitk.CopyInformation(cache["petCropped"])
            segSitk = self._keepLargestComponent(segSitk)
            newSeg = sitk.GetArrayFromImage(segSitk)

        vals = petArr[newSeg > 0]
        if len(vals) == 0:
            stats = dict(mean=0, min=0, max=0, median=0, std=0, volume_mL=0)
        else:
            stats = dict(
                mean=float(np.mean(vals)),
                min=float(np.min(vals)),
                max=float(np.max(vals)),
                median=float(np.median(vals)),
                std=float(np.std(vals)),
                volume_mL=float(len(vals)) * voxelVolMl,
            )

        cache["currentThreshold"] = float(threshold)
        cache["currentStats"] = stats
        cache["currentSegArr"] = newSeg  # cropped-space binary mask

        return stats

    def updateLivePreview(self, lesionIdx):
        """
        Push the lesion's current cropped segmentation (set by the last call
        to rethresholdLesion) into the live Slicer result segmentation node,
        so the 2D/3D views update immediately while the user drags the slider.
        """
        import SimpleITK as sitk
        import sitkUtils

        if self._resultSegNode is None or self._petSitkRef is None or self._resultLabelArr is None:
            return

        cache = self._lesionCache.get(lesionIdx)
        if cache is None or "currentSegArr" not in cache:
            return

        # Resample cropped mask back to full PET space
        segSitk = sitk.GetImageFromArray(cache["currentSegArr"].astype(np.uint8))
        segSitk.CopyInformation(cache["petCropped"])
        segFullSitk = sitk.Resample(
            segSitk, self._petSitkRef, sitk.Transform(),
            sitk.sitkNearestNeighbor, 0, sitk.sitkUInt8,
        )
        fullArr = sitk.GetArrayFromImage(segFullSitk)

        # Update only this lesion's voxels in the shared label array
        self._resultLabelArr[self._resultLabelArr == lesionIdx] = 0
        self._resultLabelArr[fullArr > 0] = lesionIdx

        self._pushLabelArrayToSegNode(self._resultLabelArr)

    def commitLesionThreshold(self, lesionIdx, resultsDf):
        """
        Finalizes the lesion's currently-selected threshold (from the last
        rethresholdLesion call) as its official result: updates resultsDf
        in place (returns a new DataFrame) and ensures the segmentation node
        reflects it. Call updateLivePreview() first, or this calls it for you.
        """
        cache = self._lesionCache.get(lesionIdx)
        if cache is None:
            raise ValueError(f"No cached data for lesion {lesionIdx}.")

        self.updateLivePreview(lesionIdx)

        stats = cache["currentStats"]
        threshold = cache["currentThreshold"]
        unit = "SUV" if self._useSUV else "kBqmL"

        threshCol = f"final_threshold_{unit}"
        meanCol = f"mean_{unit}"
        maxCol = f"max_{unit}"
        minCol = f"min_{unit}"
        medianCol = f"median_{unit}"
        stdCol = f"std_{unit}"

        mask = resultsDf["lesion_index"] == lesionIdx
        if not mask.any():
            raise ValueError(f"Lesion {lesionIdx} not found in results table.")

        resultsDf.loc[mask, "segment_name"] = cache.get("segName", f"Lesion {lesionIdx}")
        resultsDf.loc[mask, threshCol] = round(float(threshold), 4)
        resultsDf.loc[mask, "volume_mL"] = round(float(stats["volume_mL"]), 3)
        resultsDf.loc[mask, meanCol] = round(float(stats["mean"]), 4)
        resultsDf.loc[mask, maxCol] = round(float(stats["max"]), 4)
        resultsDf.loc[mask, minCol] = round(float(stats["min"]), 4)
        resultsDf.loc[mask, medianCol] = round(float(stats["median"]), 4)
        resultsDf.loc[mask, stdCol] = round(float(stats["std"]), 4)

        return resultsDf

    def _pushLabelArrayToSegNode(self, labelArr):
        """Re-import a modified label array into the existing result segmentation node."""
        import SimpleITK as sitk
        import sitkUtils

        resultLabelSitk = sitk.GetImageFromArray(labelArr)
        resultLabelSitk.CopyInformation(self._petSitkRef)
        resultLabelSitk = sitk.Cast(resultLabelSitk, sitk.sitkUInt16)

        tmpLabelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode", "__xeos_live_tmp__"
        )
        sitkUtils.PushVolumeToSlicer(resultLabelSitk, tmpLabelNode)

        # Clear and re-import into the SAME segmentation node so the viewer
        # reference stays stable (no flicker of a new node being created).
        self._resultSegNode.GetSegmentation().RemoveAllSegments()
        slicer.modules.segmentations.logic().ImportLabelmapToSegmentationNode(
            tmpLabelNode, self._resultSegNode
        )
        self._resultSegNode.CreateClosedSurfaceRepresentation()
        slicer.mrmlScene.RemoveNode(tmpLabelNode)

    # ------------------------------------------------------------------
    # Utility: crop to bounding box
    # ------------------------------------------------------------------

    @staticmethod
    def _cropToBoundingBox(petSitk, maskSitk, marginMm=5):
        import SimpleITK as sitk

        maskBin = sitk.Cast(maskSitk > 0, sitk.sitkUInt8)
        stats = sitk.LabelShapeStatisticsImageFilter()
        stats.Execute(maskBin)
        if 1 not in stats.GetLabels():
            return petSitk, maskSitk

        bb = stats.GetBoundingBox(1)  # (idx_x, idx_y, idx_z, size_x, size_y, size_z)
        spacing = petSitk.GetSpacing()
        margin_vox = [max(1, int(marginMm / s)) for s in spacing]

        start = [
            max(0, bb[i] - margin_vox[i]) for i in range(3)
        ]
        size = petSitk.GetSize()
        end_vox = [
            min(size[i] - 1, bb[i] + bb[i + 3] + margin_vox[i]) for i in range(3)
        ]
        crop_size = [end_vox[i] - start[i] for i in range(3)]

        petCropped = sitk.RegionOfInterest(petSitk, crop_size, start)
        maskCropped = sitk.RegionOfInterest(maskSitk, crop_size, start)
        return petCropped, maskCropped

    # ------------------------------------------------------------------
    # Utility: keep largest connected component
    # ------------------------------------------------------------------

    @staticmethod
    def _keepLargestComponent(segSitk):
        import SimpleITK as sitk

        segBin = sitk.Cast(segSitk > 0, sitk.sitkUInt8)
        labeled = sitk.ConnectedComponent(segBin)
        relabeled = sitk.RelabelComponent(labeled, sortByObjectSize=True)
        return sitk.Cast(relabeled == 1, sitk.sitkUInt8)

    def saveConvergencePlot(self, outputDir):
        """
        Save a single PNG with all lesion convergence curves on one figure.
        Each lesion gets its own colour. The final threshold is marked:
          ★  (star)   — threshold was manually overridden by the user
          ●  (circle) — auto-detected plateau, never changed
        Returns the path of the saved PNG.
        """
        try:
            import matplotlib
            matplotlib.use("Agg")   # headless backend — no GUI window
            import matplotlib.pyplot as plt
        except ImportError:
            slicer.util.pip_install("matplotlib")
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt

        import os

        if not self._lesionCache:
            raise ValueError("No lesion data cached. Run segmentation first.")

        unit = "SUV" if self._useSUV else "kBq/mL"
        threshCol = f"lower_threshold_{unit}"

        # Colour cycle — enough for up to 10 lesions before repeating
        colours = plt.rcParams["axes.prop_cycle"].by_key()["color"]

        fig, ax = plt.subplots(figsize=(10, 6))
        ax.set_xlabel(f"Threshold ({unit})", fontsize=12)
        ax.set_ylabel("Volume (mL)", fontsize=12)
        ax.set_title("Convergence Curves — All Lesions", fontsize=14, fontweight="bold")
        ax.grid(True, linestyle="--", alpha=0.4)

        for i, (lesionIdx, cache) in enumerate(sorted(self._lesionCache.items())):
            colour = colours[i % len(colours)]
            iterDf = cache["iterDf"]
            segName = cache.get("segName", f"Lesion {lesionIdx}")
            manualOverride = cache.get("manuallyOverridden", False)

            thresholds = iterDf[threshCol].to_numpy(dtype=float)
            volumes    = iterDf["volume_mL"].to_numpy(dtype=float)

            # Filter NaN
            valid = [(t, v) for t, v in zip(thresholds, volumes) if t == t and v == v]
            if not valid:
                continue
            tx, vy = zip(*valid)

            ax.plot(tx, vy, color=colour, linewidth=1.8, label=segName)

            # Final (committed) threshold marker
            finalThresh = cache["currentThreshold"]
            finalVol    = cache["currentStats"]["volume_mL"]

            # Find closest recorded volume to the final threshold
            closest = min(valid, key=lambda tv: abs(tv[0] - finalThresh))
            markerVol = closest[1]

            if manualOverride:
                # Star = user manually chose this threshold
                ax.plot(finalThresh, markerVol, marker="*", markersize=18,
                        color=colour, markeredgecolor="black", markeredgewidth=0.8,
                        zorder=5, linestyle="None",
                        label=f"{segName} — manual ({finalThresh:.3f})")
            else:
                # Filled circle = auto plateau
                ax.plot(finalThresh, markerVol, marker="o", markersize=11,
                        color=colour, markeredgecolor="black", markeredgewidth=0.8,
                        zorder=5, linestyle="None",
                        label=f"{segName} — auto ({finalThresh:.3f})")

        ax.legend(loc="upper right", fontsize=9, framealpha=0.85)
        fig.tight_layout()

        outPath = os.path.join(outputDir, "convergence_plot_all_lesions.png")
        fig.savefig(outPath, dpi=150)
        plt.close(fig)

        return outPath

    def generateMipGif(self, petNode, segNode, outputDir, nFrames=36, fps=12,
                       progressCallback=None):
        """
        Generate a rotating MIP GIF using the original _MIP_with_segment logic.
        Uses Pillow for GIF assembly (same as the original) and scipy for rotation.
        Returns the saved GIF path.
        """
        try:
            from PIL import Image
        except ImportError:
            slicer.util.pip_install("Pillow")
            from PIL import Image

        from scipy.ndimage import binary_erosion, gaussian_filter
        from scipy.ndimage import rotate as ndrotate
        import SimpleITK as sitk
        import sitkUtils
        import os

        # ── inner helpers (ported verbatim from _MIP_with_segment) ────

        def _normalize_to_uint8(array_2d, vmax):
            vmax = 1.0 if vmax in (None, 0) else float(vmax)
            scaled = np.clip(array_2d.astype(np.float32) / vmax, 0.0, 1.0)
            return (scaled * 255).astype(np.uint8)

        def _apply_grayscale_colormap(base_u8, cmap_name):
            gray = 255 - base_u8 if cmap_name == "gray_r" else base_u8
            return np.stack([gray, gray, gray], axis=-1).astype(np.uint8)

        def _label_palette():
            return np.array([
                [255,  64,  64], [ 64, 190, 110], [ 70, 130, 255],
                [255, 190,  60], [185,  90, 220], [ 60, 210, 220],
                [255, 120, 180], [170, 210,  70], [120, 120, 255],
                [255, 110,  50], [  0, 170, 140], [220,  70, 130],
            ], dtype=np.uint8)

        def _label_to_rgb(label_value, fallback_rgb=(255, 0, 0)):
            if label_value <= 0:
                return np.array([0, 0, 0], dtype=np.uint8)
            if int(label_value) == 1:
                return np.array(fallback_rgb, dtype=np.uint8)
            palette = _label_palette()
            return palette[(int(label_value) - 1) % len(palette)]

        def _rotate_volume(array, angle, order):
            if array is None or angle == 0:
                return array
            return ndrotate(array, angle, axes=(1, 2), reshape=False,
                            order=order, mode="constant", cval=0,
                            prefilter=(order > 1))

        def _render_surface_projection(seg_3d, fallback_rgb=(255, 0, 0), alpha_value=0.60):
            seg_3d = np.rint(seg_3d).astype(np.uint32, copy=False)
            seg_mask = seg_3d > 0
            out_shape = (seg_3d.shape[0], seg_3d.shape[2])

            if not np.any(seg_mask):
                return (np.zeros((out_shape[0], out_shape[1], 3), dtype=np.uint8),
                        np.zeros(out_shape, dtype=np.uint8))

            structure = np.zeros((3, 3, 3), dtype=bool)
            for idx in [(1,1,1),(0,1,1),(2,1,1),(1,0,1),(1,2,1),(1,1,0),(1,1,2)]:
                structure[idx] = True

            eroded = binary_erosion(seg_mask, structure=structure, border_value=0)
            surface_mask = seg_mask & (~eroded)
            has_surface = np.any(surface_mask, axis=1)
            first_surface_idx = np.argmax(surface_mask, axis=1)

            z_idx = np.arange(seg_3d.shape[0])[:, None]
            x_idx = np.arange(seg_3d.shape[2])[None, :]
            label_2d = np.zeros(out_shape, dtype=np.uint32)
            label_2d[has_surface] = seg_3d[z_idx, first_surface_idx, x_idx][has_surface]

            smooth_mask = gaussian_filter(seg_mask.astype(np.float32), sigma=1.0)
            grad_z, grad_y, grad_x = np.gradient(smooth_mask)
            normals = np.stack([-grad_z, -grad_y, -grad_x], axis=-1)
            normal_mag = np.linalg.norm(normals, axis=-1, keepdims=True)
            normals = np.divide(normals, np.maximum(normal_mag, 1e-6))

            surface_normals = np.zeros((out_shape[0], out_shape[1], 3), dtype=np.float32)
            gathered = normals[z_idx, first_surface_idx, x_idx]
            surface_normals[has_surface] = gathered[has_surface]

            thickness = np.sum(seg_mask, axis=1).astype(np.float32)
            thickness_norm = thickness / np.max(thickness) if np.max(thickness) > 0 else thickness
            depth = first_surface_idx.astype(np.float32)
            depth_norm = np.zeros_like(depth)
            if np.any(has_surface) and np.max(depth[has_surface]) > 0:
                depth_norm[has_surface] = depth[has_surface] / np.max(depth[has_surface])

            light_dir = np.array([-0.25, -0.92, 0.30], dtype=np.float32)
            light_dir /= np.linalg.norm(light_dir)
            view_dir  = np.array([0.0, -1.0, 0.0], dtype=np.float32)
            half_vec  = light_dir + view_dir
            half_vec /= np.linalg.norm(half_vec)

            diffuse  = np.clip(np.sum(surface_normals * light_dir, axis=-1), 0.0, 1.0)
            specular = np.power(np.clip(np.sum(surface_normals * half_vec, axis=-1), 0.0, 1.0), 24)
            shading  = np.clip(0.28 + 0.92*diffuse + 0.30*specular
                               + 0.18*thickness_norm - 0.10*depth_norm, 0.0, 1.0)

            rgb = np.zeros((out_shape[0], out_shape[1], 3), dtype=np.uint8)
            for label_value in np.unique(label_2d[label_2d > 0]):
                lmask = label_2d == label_value
                base_color = _label_to_rgb(int(label_value), fallback_rgb).astype(np.float32)
                lit = np.clip(base_color[None, None, :] * shading[..., None], 0, 255)
                rgb[lmask] = lit[lmask].astype(np.uint8)

            silhouette = np.zeros(out_shape, dtype=bool)
            silhouette[1:,:]  |= has_surface[1:,:]  != has_surface[:-1,:]
            silhouette[:-1,:] |= has_surface[:-1,:] != has_surface[1:,:]
            silhouette[:,1:]  |= has_surface[:,1:]  != has_surface[:,:-1]
            silhouette[:,:-1] |= has_surface[:,:-1] != has_surface[:,1:]
            silhouette &= has_surface

            label_edges = np.zeros(out_shape, dtype=bool)
            label_edges[1:,:]  |= label_2d[1:,:]  != label_2d[:-1,:]
            label_edges[:-1,:] |= label_2d[:-1,:] != label_2d[1:,:]
            label_edges[:,1:]  |= label_2d[:,1:]  != label_2d[:,:-1]
            label_edges[:,:-1] |= label_2d[:,:-1] != label_2d[:,1:]
            label_edges &= has_surface

            edge_mask = silhouette | label_edges
            rgb[edge_mask] = np.clip(rgb[edge_mask].astype(np.int16) + 35, 0, 255).astype(np.uint8)

            alpha = np.zeros(out_shape, dtype=np.uint8)
            base_alpha = np.clip(alpha_value * (0.72 + 0.28*diffuse + 0.18*thickness_norm), 0.0, 1.0)
            alpha[has_surface] = (base_alpha[has_surface] * 255).astype(np.uint8)
            alpha[edge_mask]   = np.maximum(alpha[edge_mask], 230)

            return rgb, alpha

        def _make_overlay_frame(base_2d, seg_rgb_2d=None, seg_alpha_2d=None,
                                base_vmax=None, cmap_name="gray_r"):
            base_u8  = _normalize_to_uint8(base_2d, base_vmax)
            base_rgb = _apply_grayscale_colormap(base_u8, cmap_name)
            if seg_rgb_2d is None or seg_alpha_2d is None:
                return Image.fromarray(base_rgb).convert("RGB")
            overlay_rgba = np.zeros((seg_rgb_2d.shape[0], seg_rgb_2d.shape[1], 4), dtype=np.uint8)
            overlay_rgba[..., :3] = seg_rgb_2d
            overlay_rgba[..., 3]  = seg_alpha_2d
            return Image.alpha_composite(
                Image.fromarray(base_rgb, mode="RGB").convert("RGBA"),
                Image.fromarray(overlay_rgba, mode="RGBA"),
            ).convert("RGB")

        # ── pull data from Slicer ──────────────────────────────────────

        petSitk = sitkUtils.PullVolumeFromSlicer(petNode)
        petSitk = sitk.DICOMOrient(petSitk, "LPS")

        # Export segmentation to label map in PET geometry
        tmpLabelNode = slicer.mrmlScene.AddNewNodeByClass(
            "vtkMRMLLabelMapVolumeNode", "__xeos_mip_tmp__"
        )
        try:
            slicer.modules.segmentations.logic().ExportAllSegmentsToLabelmapNode(
                segNode, tmpLabelNode, slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY
            )
            labelSitk = sitkUtils.PullVolumeFromSlicer(tmpLabelNode)
        finally:
            slicer.mrmlScene.RemoveNode(tmpLabelNode)

        labelSitk = sitk.DICOMOrient(labelSitk, "LPS")
        labelSitk = sitk.Resample(
            sitk.Cast(labelSitk, sitk.sitkUInt16),
            petSitk, sitk.Transform(),
            sitk.sitkNearestNeighbor, 0, sitk.sitkUInt16,
        )

        image_array   = sitk.GetArrayFromImage(petSitk).astype(np.float32)
        segment_array = sitk.GetArrayFromImage(labelSitk).astype(np.float32)

        # ── compute MIP frames ────────────────────────────────────────

        gif_percentile_to_show = 95
        gif_duration_ms = int(1000 / max(fps, 1))
        list_angles = list(range(0, 360, max(1, 360 // nFrames)))

        list_2d_mip_arrays      = []
        list_2d_seg_rgb_arrays  = []
        list_2d_seg_alpha_arrays = []

        for single_angle in list_angles:
            if progressCallback:
                progressCallback(len(list_2d_mip_arrays), len(list_angles))
            image_rotated   = _rotate_volume(image_array,   single_angle, 1)
            segment_rotated = _rotate_volume(segment_array, single_angle, 0)
            list_2d_mip_arrays.append(np.max(image_rotated, axis=1, keepdims=True))
            seg_rgb, seg_alpha = _render_surface_projection(segment_rotated)
            list_2d_seg_rgb_arrays.append(seg_rgb)
            list_2d_seg_alpha_arrays.append(seg_alpha)

        # ── assemble GIF (ping-pong loop, same as original) ───────────

        maximum_for_gif   = max(np.percentile(x, gif_percentile_to_show) for x in list_2d_mip_arrays)
        single_frame_shape = list_2d_mip_arrays[0].squeeze().shape
        spacing           = petSitk.GetSpacing()
        pixel_aspect_ratio = spacing[2] / spacing[0]

        gif_image_frames    = list_2d_mip_arrays    + list_2d_mip_arrays[::-1]
        gif_seg_rgb_frames  = list_2d_seg_rgb_arrays  + list_2d_seg_rgb_arrays[::-1]
        gif_seg_alpha_frames = list_2d_seg_alpha_arrays + list_2d_seg_alpha_arrays[::-1]

        frames = []
        for idx, arr in enumerate(gif_image_frames):
            arr2d      = np.flip(np.squeeze(arr), 0)
            seg_rgb_2d = np.flip(gif_seg_rgb_frames[idx],   0)
            seg_alpha_2d = np.flip(gif_seg_alpha_frames[idx], 0)

            pil_img = _make_overlay_frame(
                arr2d,
                seg_rgb_2d=seg_rgb_2d,
                seg_alpha_2d=seg_alpha_2d,
                base_vmax=maximum_for_gif,
                cmap_name="gray_r",
            )

            if abs(pixel_aspect_ratio - 1.0) > 0.1:
                target_width  = single_frame_shape[1]
                target_height = max(1, int(single_frame_shape[0] * pixel_aspect_ratio))
                pil_img = pil_img.resize((target_width, target_height), Image.LANCZOS)

            frames.append(pil_img.convert("P", palette=Image.ADAPTIVE, colors=256))

        outPath = os.path.join(outputDir, "MIP_rotating.gif")
        if frames:
            frames[0].save(
                outPath,
                format="GIF",
                append_images=frames[1:],
                save_all=True,
                duration=gif_duration_ms,
                loop=1,
            )

        return outPath

    def exportImages(self, exportItems, outputDir, fmt="nrrd"):
        """
        Export volumes and segmentation nodes to files.

        exportItems: list of (label, node, isSegmentation)
            label          — filename stem (e.g. "PET", "CT", "FinalSegmentation")
            node           — vtkMRMLScalarVolumeNode or vtkMRMLSegmentationNode
            isSegmentation — if True, export via label-map conversion
        fmt: "nrrd" | "nii.gz" | "mha"

        Returns list of saved absolute paths.
        """
        import SimpleITK as sitk
        import sitkUtils
        import os

        os.makedirs(outputDir, exist_ok=True)
        savedPaths = []

        for label, node, isSegmentation in exportItems:
            if node is None:
                continue

            if isSegmentation:
                # Convert segmentation → label map volume → SimpleITK → file
                tmpLabelNode = slicer.mrmlScene.AddNewNodeByClass(
                    "vtkMRMLLabelMapVolumeNode", "__xeos_export_tmp__"
                )
                try:
                    # Export all visible segments into the label map
                    slicer.modules.segmentations.logic().ExportAllSegmentsToLabelmapNode(
                        node, tmpLabelNode, slicer.vtkSegmentation.EXTENT_REFERENCE_GEOMETRY
                    )
                    sitkImg = sitkUtils.PullVolumeFromSlicer(tmpLabelNode)
                    sitkImg = sitk.Cast(sitkImg, sitk.sitkUInt16)
                    outPath = os.path.join(outputDir, f"{label}.{fmt}")
                    sitk.WriteImage(sitkImg, outPath, useCompression=True)
                    savedPaths.append(outPath)
                finally:
                    slicer.mrmlScene.RemoveNode(tmpLabelNode)

            else:
                # Scalar volume — pull and write directly
                sitkImg = sitkUtils.PullVolumeFromSlicer(node)
                outPath = os.path.join(outputDir, f"{label}.{fmt}")
                sitk.WriteImage(sitkImg, outPath, useCompression=True)
                savedPaths.append(outPath)

        return savedPaths

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def exportToExcel(self, resultsDf, lesionResultDfs, outputDir):
        """
        Saves:
          - iterative-segmentation-summary.xlsx  (one row per lesion)
          - iterative-segmentation-lesion-<N>.xlsx  (per-lesion iterations)
        Returns list of saved file paths.
        """
        try:
            import openpyxl
        except ImportError:
            slicer.util.pip_install("openpyxl")
            import openpyxl
        try:
            import pandas as pd
        except ImportError:
            slicer.util.pip_install("pandas")
            import pandas as pd

        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows

        savedPaths = []

        def _style_sheet(ws, df, title):
            header_font = Font(name="Arial", bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="2E4057")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
            ws.title = title[:31]

        # Summary sheet
        summaryPath = os.path.join(outputDir, "iterative-segmentation-summary.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in dataframe_to_rows(resultsDf, index=False, header=True):
            ws.append(r)
        _style_sheet(ws, resultsDf, "Summary")
        wb.save(summaryPath)
        savedPaths.append(summaryPath)

        # Per-lesion iteration sheets in one workbook
        detailPath = os.path.join(outputDir, "iterative-segmentation-iterations.xlsx")
        wbD = openpyxl.Workbook()
        wbD.remove(wbD.active)  # remove default sheet
        for lesionIdx, df in lesionResultDfs.items():
            wsD = wbD.create_sheet(title=f"Lesion_{lesionIdx}")
            for r in dataframe_to_rows(df, index=False, header=True):
                wsD.append(r)
            _style_sheet(wsD, df, f"Lesion_{lesionIdx}")
        wbD.save(detailPath)
        savedPaths.append(detailPath)

        return savedPaths


# ─────────────────────────────────────────────────────────────────────────────
# Helper shim: sitk.GetArrayFromImage may need import guard
# ─────────────────────────────────────────────────────────────────────────────
def sitk_GetArrayFromImage(sitkImage):
    import SimpleITK as sitk
    return sitk.GetArrayFromImage(sitkImage)


# ─────────────────────────────────────────────────────────────────────────────
# Test stub (required by Slicer convention)
# ─────────────────────────────────────────────────────────────────────────────
class XEOSIterativeSegmentationTest(ScriptedLoadableModuleTest):
    def setUp(self):
        slicer.mrmlScene.Clear()

    def runTest(self):
        self.setUp()
        self.test_placeholder()

    def test_placeholder(self):
        self.delayDisplay("No automated tests configured yet.")
